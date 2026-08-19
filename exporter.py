from __future__ import annotations

import io
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def _wa_link(phone: str) -> str:
    """Convert phone number to wa.me link. Returns '' if phone empty."""
    if not phone or phone == "-":
        return ""
    # Strip non-digit chars except leading +
    digits = re.sub(r"[^\d]", "", phone)
    # Indonesian: leading 0 → 62
    if digits.startswith("0"):
        digits = "62" + digits[1:]
    return f"https://wa.me/{digits}"


def build_excel(data: list[dict], keyword: str) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Hasil Scraping"

    headers = ["Nama", "No. Telepon", "Link WhatsApp", "Email", "Link Google Maps", "Website"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1A73E8")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 20

    for row_idx, item in enumerate(data, 2):
        phone = item.get("no_telepon", "") or "-"

        ws.cell(row=row_idx, column=1, value=item.get("nama", "") or "-")
        ws.cell(row=row_idx, column=2, value=phone)

        # WhatsApp link
        wa_url = _wa_link(phone)
        cell = ws.cell(row=row_idx, column=3, value=wa_url if wa_url else "-")
        if wa_url:
            cell.hyperlink = wa_url
            cell.font = Font(color="25D366", underline="single")

        ws.cell(row=row_idx, column=4, value=item.get("email", "") or "-")

        maps_url = item.get("link_google_maps", "")
        cell = ws.cell(row=row_idx, column=5, value=maps_url or "-")
        if maps_url:
            cell.hyperlink = maps_url
            cell.font = Font(color="1A73E8", underline="single")

        website = item.get("website", "")
        cell = ws.cell(row=row_idx, column=6, value=website or "-")
        if website:
            cell.hyperlink = website
            cell.font = Font(color="1A73E8", underline="single")

    col_widths = [40, 20, 35, 30, 50, 40]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
